"""Import loans and payments from the Loans Excel file."""
from __future__ import annotations
import warnings
from pathlib import Path
from datetime import datetime

import openpyxl

import db

warnings.filterwarnings("ignore")

LOANS_EXCEL = Path(r"Z:\salaries\Loans.xlsb.xlsx")


def _date(v) -> str:
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str) and v.strip():
        return v.strip()[:10]
    return ""


def _f(v) -> float:
    try:
        return float(v) if v else 0.0
    except (TypeError, ValueError):
        return 0.0


def _s(v) -> str:
    return str(v).strip() if v is not None else ""


def import_loans(path: Path = LOANS_EXCEL) -> tuple[int, int]:
    """Import all loans and payments. Returns (loans_count, payments_count)."""
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    # ── Collect loans ────────────────────────────────────────────────────────
    loans: list[dict] = []

    # Loans sheet: cols 0-5 → Loan_No, staff_ID, Date, Amount, Pmt_Amount, Name
    ws = wb["Loans"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[2]:
            continue
        amt = _f(row[3])
        if amt <= 0:
            continue
        loans.append({
            "loan_no":    int(_f(row[0])),
            "staff_id":   _s(row[1]),
            "date":       _date(row[2]),
            "amount":     amt,
            "pmt_amount": _f(row[4]),
            "name":       _s(row[5]),
        })

    # ── Collect payments ─────────────────────────────────────────────────────
    # Only Payments sheet is authoritative — Sheet2 loans/payments are duplicates
    pays: list[dict] = []

    wsp = wb["Payments"]
    for row in wsp.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[2] or not row[4]:
            continue
        pays.append({
            "date":     _date(row[0]),
            "pay_no":   int(_f(row[1])),
            "name":     _s(row[2]),
            "staff_id": _s(row[3]),
            "amount":   _f(row[4]),
        })

    wb.close()

    # Deduplicate by (name, date, amount) to avoid double-import
    seen_loans: set = set()
    seen_pays:  set = set()

    db.init_db()

    loans_added = 0
    for l in loans:
        key = (l["loan_no"], l["name"], l["date"])
        if key in seen_loans or not l["date"] or not l["name"]:
            continue
        seen_loans.add(key)
        db.add_loan(
            employee_name=l["name"],
            loan_date=l["date"],
            principal=l["amount"],
            pmt_amount=l["pmt_amount"],
            staff_id=l["staff_id"],
            loan_no=l["loan_no"],
        )
        loans_added += 1

    pays_added = 0
    for p in pays:
        key = (p["pay_no"], p["name"], p["date"], p["amount"])
        if key in seen_pays or not p["date"] or not p["name"] or p["amount"] <= 0:
            continue
        seen_pays.add(key)
        db.add_loan_payment(
            employee_name=p["name"],
            payment_date=p["date"],
            amount=p["amount"],
            staff_id=p["staff_id"],
            pay_no=p["pay_no"],
        )
        pays_added += 1

    return loans_added, pays_added


if __name__ == "__main__":
    l, p = import_loans()
    print(f"Imported {l} loans and {p} payments.")
    for name in db.get_loan_names():
        s = db.get_loan_summary(name)
        print(f"  {name:<30} loaned=${s['total_loaned']:>9,.2f}  "
              f"paid=${s['total_paid']:>9,.2f}  balance=${s['balance']:>9,.2f}")
